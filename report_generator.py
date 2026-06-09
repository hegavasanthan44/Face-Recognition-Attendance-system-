from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from config import REPORTS_DIR
from attendance_manager import AttendanceManager

class ReportGenerator:
    def __init__(self):
        self.attendance_manager = AttendanceManager()

    def export_attendance_to_excel(self, filename=None):
        if not filename:
            filename = f"attendance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        filepath = REPORTS_DIR / filename

        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"

        self._add_attendance_sheet(ws)

        wb.save(str(filepath))
        return filepath

    def _add_attendance_sheet(self, ws):
        stats = self.attendance_manager.get_all_statistics()

        headers = ["Name", "Total Days", "Present", "Late", "Attendance %"]
        ws.append(headers)

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for stat in stats:
            ws.append([
                stat["name"],
                stat["total_days"],
                stat["present"],
                stat["late"],
                f"{stat['percentage']}%"
            ])

        summary_row = len(stats) + 3
        ws[f"A{summary_row}"] = "Summary"
        ws[f"A{summary_row}"].font = Font(bold=True, size=12)

        summary = self.attendance_manager.get_attendance_summary()
        if summary:
            ws[f"A{summary_row + 1}"] = "Total Registered:"
            ws[f"B{summary_row + 1}"] = summary["total_registered"]

            ws[f"A{summary_row + 2}"] = "Average Attendance:"
            ws[f"B{summary_row + 2}"] = f"{summary['average_attendance']}%"

        for col in ["A", "B", "C", "D", "E"]:
            ws.column_dimensions[col].width = 15

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for row in ws.iter_rows(min_row=1, max_row=len(stats) + 1, min_col=1, max_col=5):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:
                    cell.alignment = Alignment(horizontal="center")

    def generate_analytics_report(self):
        print("\n" + "="*60)
        print("ATTENDANCE ANALYTICS REPORT")
        print("="*60)
        print(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

        summary = self.attendance_manager.get_attendance_summary()

        if not summary:
            print("No attendance data available.")
            return

        print(f"Total Registered Personnel: {summary['total_registered']}")
        print(f"Average Attendance Rate: {summary['average_attendance']}%\n")

        print("-" * 60)
        print(f"{'Name':<20} {'Total':<8} {'Present':<10} {'Late':<8} {'%':<8}")
        print("-" * 60)

        for stat in summary["details"]:
            print(f"{stat['name']:<20} {stat['total_days']:<8} {stat['present']:<10} "
                  f"{stat['late']:<8} {stat['percentage']:<8.2f}%")

        print("-" * 60)
